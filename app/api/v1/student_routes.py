from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from app.db.database import get_database_session
from app.models.models import Student, Class
from app.api.v1.validate import validate_id  # Import the generic ID validation function
from app.schemas.student_schema import StudentCreate, StudentUpdate, StudentResponse
from typing import List
import openpyxl
import pandas as pd


router = APIRouter(prefix="/students", tags=["Students"])


@router.get("/get_all/", response_model=List[StudentResponse])
def get_all_student(db: Session = Depends(get_database_session),
                 offset: int = 0,
                 limit: int = 10):
    """
    Fetch students with optional search term and pagination.
    """
    query = db.query(Student)
    return query.offset(offset).limit(limit).all()


@router.get("/get_ID/", response_model=List[StudentResponse])
def get_student_id(db: Session = Depends(get_database_session),
                 search_student_id: str = None,
                 offset: int = 0,
                 limit: int = 10):
    """
    Fetch students with optional search term and pagination.
    """
    query = db.query(Student)
    if search_student_id:
        query = query.filter(Student.student_ID.contains(search_student_id))
    return query.offset(offset).limit(limit).all()


@router.get("/get_name/", response_model=List[StudentResponse])
def get_student_name(db: Session = Depends(get_database_session),
                 search_student_name: str = None,
                 offset: int = 0,
                 limit: int = 10):
    """
    Fetch students with optional search term and pagination.
    """
    query = db.query(Student)
    if search_student_name:
        query = query.filter(Student.student_name.contains(search_student_name))
    return query.offset(offset).limit(limit).all()


@router.post("/", response_model=StudentResponse)
def create_student(student_data: StudentCreate,
                   db: Session = Depends(get_database_session)):
    """
    Create a new student.
    """
    new_student = Student(**student_data)
    db.add(new_student)
    db.commit()
    db.refresh(new_student)
    return new_student


@router.put("/{student_id}")
def update_student(student_id: int,
                   student_data: dict,
                   db: Session = Depends(get_database_session)):
    """
    Update an existing student.
    """
    # Validate and fetch the student instance
    student_instance = validate_id(Student, student_id, db, id_field="student_id")
    for key, value in student_data.items():
        setattr(student_instance, key, value)
    db.commit()
    return student_instance

@router.delete("/{student_id}")
def delete_student(student_id: int,
                   db: Session = Depends(get_database_session)):
    """
    Delete a student.
    """
    # Validate and fetch the student instance
    student_instance = validate_id(Student, student_id, db, id_field="student_id")
    db.delete(student_instance)
    db.commit()
    return {"detail": "Student deleted successfully"}



@router.post("/upload/")
def upload_students(file: UploadFile = File(...), db: Session = Depends(get_database_session)):
    """
    Upload students from Excel or CSV file.
    """
    try:
        # Determine file type
        if file.filename.endswith(".xlsx"):
            file.file.seek(0)  # Reset file pointer
            workbook = openpyxl.load_workbook(file.file)
            worksheet = workbook.active

            # Validate structure
            if worksheet.max_column < 3:
                raise HTTPException(status_code=400, detail="Invalid Excel file structure. Expected columns: student_number, student_name, class_name.")

            data = [
                (row[0], row[1], row[2])
                for row in worksheet.iter_rows(min_row=2, values_only=True)
            ]

        elif file.filename.endswith(".csv"):
            file.file.seek(0)  # Reset file pointer
            df = pd.read_csv(file.file)

            # Validate structure
            expected_columns = {"student_number", "student_name", "class_name"}
            if expected_columns.difference(df.columns):
                raise HTTPException(status_code=400, detail=f"Invalid CSV file structure. Expected columns: {', '.join(expected_columns)}.")

            data = df[["student_number", "student_name", "class_name"]].values.tolist()

        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Upload an .xlsx or .csv file.")

        # Process rows
        for student_number, student_name, class_name in data:
            if not student_number or not student_name or not class_name:
                raise HTTPException(status_code=400, detail="Missing required fields: student_number, student_name, or class_name.")

            # Check class existence
            class_instance = db.query(Class).filter(Class.class_name == class_name).first()
            if not class_instance:
                raise HTTPException(status_code=404, detail=f"Class with name '{class_name}' not found.")

            # Check if student already exists
            student_exists = db.query(Student).filter(Student.student_number == student_number).first()
            if student_exists:
                continue  # Skip duplicate students

            # Create new student
            new_student = Student(
                student_number=student_number,
                student_name=student_name,
                class_id=class_instance.class_id
            )
            db.add(new_student)

        db.commit()
        return {"detail": "Students uploaded successfully"}

    except openpyxl.utils.exceptions.InvalidFileException:
        raise HTTPException(status_code=400, detail="Invalid Excel file format. Please upload a valid .xlsx file.")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Invalid CSV file format. Please upload a valid .csv file.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unexpected error: {str(e)}")
